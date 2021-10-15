#!/usr/bin/env python3
# coding: utf-8
import csv
import os
import random
import re
import time
import traceback
import openpyxl.utils.exceptions
import openpyxl
import xlrd
import json
import logging
import pymongo
import datetime
import sqlite3
import pandas as pd
import copy
import gridfs
import abc
from bson import ObjectId
import pymongo.errors
# import bson.objectid
from abc import ABC
from pathlib import Path
from .myutils import EsModel
from .myutils import BaseDB
from .myutils import ClientPyMySQL
from .myutils import ExcelWriter
from .sqlfileextra import SqlExtractor, match_insert, RANDOM_STR
from ..utils import format_error, with_cur_lock, gen_pass, run_task_auto_retry
from clickhouse_driver import connect as clickhouse_connect
from threading import Lock
from threading import local as threading_local


def get_realpath():
    return os.path.split(os.path.realpath(__file__))[0]


def secure_filename(filename):
    if isinstance(filename, str):
        from unicodedata import normalize
        filename = normalize('NFKD', filename).encode('utf-8', 'ignore').decode('utf-8')  # 转码

    for sep in os.path.sep, os.path.altsep:
        if sep:
            filename = filename.replace(sep, ' ')

    # 正则增加对汉字和日语假名的部分（本人有需求）
    # \ / : * ? " < > |    r"[\/\\\:\*\?\"\<\>\|]"
    # [ \\[ \\] \\^ \\-_*×――(^)$%~!@#$…&%￥—+=<>《》!！??？:：•`·、。，；,.;\"‘’“”-]
    # \u4E00-\u9FBF 中文
    # \u3040-\u30FF 假名
    # \u31F0-\u31FF 片假名扩展
    # _filename_ascii_add_strip_re = re.compile(r'[^A-Za-z0-9_\u4E00-\u9FBF\u3040-\u30FF\u31F0-\u31FF.-]')
    _filename_ascii_add_strip_re = re.compile(r"[\/\\\:\*\?\"\<\>\|/]")

    filename = str(_filename_ascii_add_strip_re.sub('', '_'.join(  # 新的正则
        filename.split()))).strip('._')

    _windows_device_files = ('CON', 'AUX', 'COM1', 'COM2', 'COM3', 'COM4', 'LPT1',
                             'LPT2', 'LPT3', 'PRN', 'NUL')

    # on nt a couple of special files are present in each folder.  We
    # have to ensure that the target file is not such a filename.  In
    # this case we prepend an underline
    if os.name == 'nt' and filename and \
            filename.split('.')[0].upper() in _windows_device_files:
        filename = '_' + filename

    return filename[:250]


class ElasticSearchD(EsModel):
    def __init__(self, hosts, username=None, password=None, cols_ddl=None, ik=False):
        super().__init__(hosts, username, password)
        self.hosts = hosts
        self.username = username
        self.password = password
        self.ik = ik
        self.cols_ddl = cols_ddl

    def __repr__(self):
        return f'ElasticSearch:{self.hosts}'

    def get_data(self, index, *args, **kwargs):
        if isinstance(index, str):
            if 'condition' in kwargs:
                query = kwargs.pop('condition')
            else:
                query = {}
        else:
            query = index[1]
            index = index[0]
        logging.debug(json.dumps(query))
        for i in self.scan(query=query, index=index, *args, **kwargs):
            r = {}
            if r.get('id'):
                r['_id'] = i['_id']
            else:
                r['id'] = i['_id']
            r.update(i['_source'])
            yield r

    def save_data(self, index, data, batch_size=1000, pks='', pop_id='_id', *args, **kwargs):
        actions = []
        write_kws = {k: kwargs[k] for k in ['batch_size', 'retry'] if k in kwargs}
        for d in data:
            if pks:
                _id = '-'.join(d[k] for k in pks.split(','))
            elif '_id' in d:
                _id = d['_id']
            else:
                _id = '_'.join(map(lambda x: str(x) if len(str(x)) <= 10 else str(x)[0:11], d.values()))

            if pop_id:
                for k in pop_id.split(','):
                    d.pop(k)
            actions.append({
                '_op_type': 'index',  # 操作 index update create delete
                '_index': index,  # index
                '_type': '_doc',  # type
                "_id": f"{_id}",
                '_source': d})
            if len(actions) > batch_size:
                self.bulk_write(actions, **write_kws)
                actions = []

        if len(actions):
            self.bulk_write(actions, **write_kws)


    def get_indexes(self):
        return list(self.es.indices.get_alias().keys())

    def get_ddl(self, index):
        return self.es.indices.get_mapping(index)

    def get_count(self, index, *args, **kwargs):
        if isinstance(index, str):
            query = {}
        else:
            query = {'query': index[1]['query']}
            index = index[0]
            if query.get("_source"):
                query.pop("_source")
        return int(self.es.count(index=index, body=query)['count'])

    @staticmethod
    def get_int_type_from_len(length):
        if 0 < length <= 8:
            return "byte"
        elif 8 < length <= 16:
            return "short"
        elif 16 < length <= 32:
            return "integer"
        elif 32 < length <= 64:
            return "long"
        return None

    @staticmethod
    def get_str_type_from_len(length):
        if 0 < length <= 256:
            return "keyword"
        elif 256 < length <= 65536:
            return "text"
        elif 65536 < length <= 16777216:
            return "text"
        elif 16777216 < length <= 4294967295:
            return "text"
        return None

    def create_index(self, index, data, pks='id'):
        # sql = """SET NAMES utf8mb4;"""
        # sql += """SET FOREIGN_KEY_CHECKS = 0;"""
        # if drop == 1:
        #     sql += """DROP TABLE IF EXISTS `{}`;""".format(tbname)
        pairs = {
            "settings": {
                "index": {
                    "number_of_shards": 1,
                    "number_of_replicas": 1
                }
            },
            'mappings': {"doc": {
                'properties': {

                }}
            }
        }
        properties = pairs['mappings']['doc']['properties']
        # ES7以上删掉doc层
        # pairs = {
        #     "settings": {
        #         "index": {
        #             "number_of_shards": 1,
        #             "number_of_replicas": 1
        #         }
        #     },
        #     'mappings': {
        #         'properties': {
        #
        #         }
        #     }
        # }
        # properties = pairs['mappings']['properties']

        if self.cols_ddl:
            if self.cols_ddl[index].get('mappings'):
                pairs['mappings'] = self.cols_ddl[index]['mappings']
            else:
                for col, col_type in self.cols_ddl[index].items():
                    properties[col] = {}
                    if col_type in ['int', 'bigint', 'integer', 'tinyint', 'smallint', 'mediumint']:
                        properties[col]['type'] = 'long'
                    elif col_type in ['float', 'double']:
                        properties[col]['type'] = 'float'
                    elif col_type in ['bit']:
                        properties[col]['type'] = 'long'
                    elif col_type in ['char', 'varchar', 'text', 'tinyblob', 'tinytext', 'blob', 'mediumtext', 'mediumblob',
                                      'longtext', 'longblob', 'json', 'timestamp', 'datetime']:
                        properties[col] = {
                            "type": "text",
                            "fields": {
                                "keyword": {
                                    "type": "keyword",
                                    "ignore_above": 1024
                                }
                            },
                        }
                        if self.ik:
                            properties[col]["fields"]["analyzer"] = "ik_max_word"
                            properties[col]["fields"]["search_analyzer"] = "ik_smart"
        else:
            for name, value in data.items():
                if isinstance(value, dict):
                    continue
                properties[name] = {}
                if isinstance(value, int):
                    properties[name]['type'] = 'long'
                elif isinstance(value, float):
                    properties[name]['type'] = 'float'
                elif isinstance(value, str) or isinstance(value, list):
                    properties[name] = {
                        "type": "text",
                        "fields": {
                            "keyword": {
                                "type": "keyword",
                                "ignore_above": 1024
                            }
                        },
                    }
                    if self.ik:
                        properties[name]["fields"]["analyzer"] = "ik_max_word"
                        properties[name]["fields"]["search_analyzer"] = "ik_smart"
        if self.es.indices.exists(index=index) is not True:
            res = self.es.indices.create(index=index, body=pairs)
            if not res:
                logging.error("错误，创建索引{}失败".format(index))
                return False

            logging.info("正常，创建索引{}成功".format(index))
            return True
        else:
            logging.info("正常，索引{}已存在".format(index))
            return True


# class MySqlD(BaseDB, ABC):
#     def __init__(self, host, port, user, passwd, database=None):
#         super().__init__(host, port, database, user, passwd)
#
#     def get_data(self, index, *args, **kwargs):
#         return self.__execute__(sql=f'select * from {index}', *args, **kwargs).fetchall()
#
#     def save_data(self, index, *args, **kwargs):
#         self.insert_many_with_dict_list(tablename=index, *args, **kwargs)
#
#     def get_indexes(self):
#         tbs = []
#         for r in list(self.__execute__(sql='show tables;')):
#             tbs.append(r[0])
#         return tbs
#
#     def get_count(self, index):
#         return int(self.__execute__(sql=f'select count(0) as c from {index}').fetchone()[0])
#
#     @staticmethod
#     def get_int_type_from_len(length):
#         if 0 < length <= 8:
#             return "tinyint"
#         elif 8 < length <= 16:
#             return "smallint"
#         elif 16 < length <= 32:
#             return "int"
#         elif 32 < length <= 64:
#             return "bigint"
#         return None
#
#     @staticmethod
#     def get_str_type_from_len(length):
#         if 0 < length <= 256:
#             return "varchar({})".format(length)
#         elif 256 < length <= 65536:
#             return "text"
#         elif 65536 < length <= 16777216:
#             return "mediumtext"
#         elif 16777216 < length <= 4294967295:
#             return "longtext"
#         return None
#
#     def create_index(self, index, data, pks='id'):
#         # sql = """SET NAMES utf8mb4;"""
#         # sql += """SET FOREIGN_KEY_CHECKS = 0;"""
#         # if drop == 1:
#         #     sql += """DROP TABLE IF EXISTS `{}`;""".format(tbname)
#         sql = """create table if not exists {}(""".format(index)
#         for name, value in data.items():
#             if isinstance(value, int):
#                 _type = 'bigint'
#             elif isinstance(value, str):
#                 if pks.find(name) > -1:
#                     _type = 'varchar(256)'
#                 else:
#                     _type = 'text'
#             else:
#                 _type = 'blob'
#
#             sql += ('`' + name + '` ')
#             sql += (_type + ',')
#
#         slen = len(pks)
#         if slen > 0:
#             sql += """PRIMARY KEY ("""
#             v_keys = pks.split(',')
#             first = 1
#             for key in v_keys:
#                 if first == 1:
#                     first = 0
#                     sql += """`{}`""".format(key)
#                 else:
#                     sql += """,`{}`""".format(key)
#             sql += """)) """
#         else:
#             sql = sql[0: -1]
#             sql += ')'
#
#         sql += """ENGINE = InnoDB CHARACTER SET = utf8mb4 COLLATE = utf8mb4_general_ci ROW_FORMAT = Dynamic;"""
#         print(sql)
#         self.__execute__(sql)
#         self.__commit__()


class MySqlD(ClientPyMySQL, ABC):
    def __init__(self, host, port, user, passwd, database=None, tables_ddl=None):
        super().__init__(host, port, database, user, passwd)
        self.host = host
        self.port = port
        self.database = database
        self.tables_ddl = tables_ddl

    def __repr__(self):
        return f'MySQL:{self.host}:{self.port}/{self.database}'

    def get_data(self, index, *args, **kwargs):
        sub_sql = f"{index} {kwargs['condition']}" if 'condition' in kwargs else index
        if index.lower().strip().startswith('select '):
            return self._execute(sql=f'{sub_sql}', *args, **kwargs)[1]
        else:
            return self._execute(sql=f'select * from {sub_sql}', *args, **kwargs)[1]

    def save_data(self, index, data, *args, **kwargs):
        self.insert_many_with_dict_list(tablename=index, data=data, *args, **kwargs)

    def get_indexes(self):
        return list(t[f'Tables_in_{self.database}'] for t in self._execute(sql='show tables;')[1])

    def get_columns_type(self, index):
        sql = f"select table_name,column_name,data_type,column_comment,column_default,column_key " \
              f"from information_schema.columns " \
              f"where table_schema='{self.database}' and table_name='{index}';"
        return list(t for t in self._execute(sql=sql)[1])

    def get_table_ddl(self, index):
        sql = f"show create table {index}"
        return 'CREATE TABLE IF NOT EXISTS ' + self._execute(sql=sql)[1].__next__()['Create Table'].replace('\n', '').strip('CREATE TABLE')

    def get_count(self, index, *args, **kwargs):
        if index.lower().strip().startswith('select '):
            index = f'({index.lower().strip()})'
        for r in self._execute(f'select count(0) as c from  {index} as taSDFEWVempTABlesdfecH', )[1]:
            return r['c']

    @staticmethod
    def get_int_type_from_len(length):
        if 0 < length <= 8:
            return "tinyint"
        elif 8 < length <= 16:
            return "smallint"
        elif 16 < length <= 32:
            return "int"
        elif 32 < length <= 64:
            return "bigint"
        return None

    @staticmethod
    def get_str_type_from_len(length):
        if 0 < length <= 256:
            return "varchar({})".format(length)
        elif 256 < length <= 65536:
            return "text"
        elif 65536 < length <= 16777216:
            return "mediumtext"
        elif 16777216 < length <= 4294967295:
            return "longtext"
        return None

    def create_index(self, index, data, pks='id'):
        if self.tables_ddl:
            sql = self.tables_ddl[index]
            self._execute(sql)
            self.end_transaction()
        else:
            # sql = """SET NAMES utf8mb4;"""
            # sql += """SET FOREIGN_KEY_CHECKS = 0;"""
            # if drop == 1:
            #     sql += """DROP TABLE IF EXISTS `{}`;""".format(tbname)
            sql = """create table if not exists {}(""".format(index)
            for name, value in data.items():
                if isinstance(value, int):
                    _type = 'bigint'
                elif isinstance(value, str):
                    if name in pks.split(','):
                        _type = 'varchar(256)'
                    else:
                        _type = 'text'
                else:
                    _type = 'blob'

                sql += ('`' + name + '` ')
                sql += (_type + ',')

            slen = len(pks)
            if slen > 0:
                sql += """PRIMARY KEY ("""
                v_keys = pks.split(',')
                first = 1
                for key in v_keys:
                    if first == 1:
                        first = 0
                        sql += """`{}`""".format(key)
                    else:
                        sql += """,`{}`""".format(key)
                sql += """)) """
            else:
                sql = sql[0: -1]
                sql += ')'

            sql += """ENGINE = InnoDB CHARACTER SET = utf8mb4 COLLATE = utf8mb4_general_ci ROW_FORMAT = Dynamic;"""
            print(sql)
            self._execute(sql)
            self.end_transaction()


def get_line_num_fast(filename):
    count_n = 0
    count_r = 0
    fp = open(filename, "rb")
    while 1:
        buffer = fp.read(1 * 1024 * 1024)
        if not buffer:
            break
        count_n += buffer.count(b'\n')
        count_r += buffer.count(b'\r')
    fp.close()
    return count_n or count_r


class BaseFileD(object):
    def __init__(self, path, extension, encoding='utf8-sig', newline=None):
        self.path = path
        self.extension = extension
        self.encoding = encoding
        self.newline = newline
        self._file_w = dict()
        self._indexes_path = dict()

    def __repr__(self):
        return f'{self.extension}:{self.path}'

    def __del__(self):
        for f in self._file_w.values():
            try:
                f.close()
            except Exception as e:
                logging.warning(e)

    def gen_path_by_index(self, index):
        if index not in self._indexes_path:
            path = f'{self.path}{os.sep}{secure_filename(index)}.{self.extension}'
            self._indexes_path[index] = path
        return self._indexes_path[index]

    def get_indexes(self):
        res = []
        real_path = os.path.split(os.path.realpath(self.path))[0]
        for root, fs, fns in os.walk(self.path):
            for fn in fns:
                path = f'{root}{os.pathsep}{fn}'
                (filepath, tempfilename) = os.path.split(path)
                (filename, extension) = os.path.splitext(tempfilename)
                if extension not in [f'.{self.extension}']:
                    continue
                index = '-'.join(list(filepath.replace(real_path, '', 1).split(os.sep)) + [filename])
                res.append(index)
                self._indexes_path[index] = path
        return res

    def get_count(self, index, *args, **kwargs):
        return get_line_num_fast(self.gen_path_by_index(index))

    @classmethod
    def w_open_func(cls, *args, **kwargs):
        return open(*args, **kwargs)

    def create_index(self, index, data, pks='id'):
        if not os.path.exists(self.path):
            try:
                os.makedirs(self.path)
            except Exception as e:
                logging.warning(e)
                Path(self.path).mkdir(parents=True, exist_ok=True)
        path = self.gen_path_by_index(index)
        if os.path.exists(path):
            os.rename(path, f"{path}.{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.bak")
        dirname = os.path.dirname(path)
        if not os.path.exists(dirname):
            os.makedirs(dirname)
        self._file_w.setdefault(index, self.w_open_func(path, 'w', encoding=self.encoding, newline=self.newline))


class CsvD(BaseFileD):
    def __init__(self, path, split=',', extension='csv', encoding='utf8'):
        super().__init__(path, extension, encoding, newline='')
        self.split = split
        self._file_w = dict()
        self.___file_w = dict()

    def get_count(self, index, *args, **kwargs):
        count = get_line_num_fast(self.gen_path_by_index(index))
        if count:
            count -= 1
        return count

    @property
    def __file_w(self):
        for k, v in self._file_w.items():
            if k not in self.___file_w:
                self.___file_w[k] = csv.writer(v)
        return self.___file_w

    # @classmethod
    # def w_open_func(cls, *args, **kwargs):
    #     return csv.writer(open(*args, **kwargs))

    def get_data(self, index, fieldnames=None, restkey=None, restval=None,
                 dialect="excel", **kwargs):
        with open(self.gen_path_by_index(index), 'r', encoding=self.encoding) as f:
            for line in csv.DictReader(
                    f, fieldnames=fieldnames, restkey=restkey, restval=restval, dialect=dialect, **kwargs):
                yield line
            # keys = list(k.replace("'", '').replace('"', '') for k in f.readline().strip().split(self.split))
            # for line in f:
            #     yield {keys[idx]: v for idx, v in enumerate(eval(line))}

    def save_data(self, index, data, *args, **kwargs):
        # self._file_w[index].writelines((self.split.join(f'{v.__repr__()}' for v in d.values()) + '\n') for d in data)
        self.__file_w[index].writerows([v for v in d.values()] for d in data)
        self._file_w[index].flush()

    def create_index(self, index, data, pks='id'):
        super(self.__class__, self).create_index(index, data)
        # self._file_w[index].writerow((self.split.join(f'"{v.__repr__()[1:-1]}"' for v in data.keys()) + '\n'))
        self.__file_w[index].writerow([v for v in data.keys()])
        self._file_w[index].flush()


class SqlFileD(BaseFileD):
    def __init__(self, path, extension='sql', encoding='utf8', mode='insert', compress=False):
        super().__init__(path, extension, encoding)
        self.compress = compress
        self.mode = {
            1: 'INSERT',
            "insert": 'INSERT',
            "INSERT": 'INSERT',
            2: 'INSERT IGNORE',
            'insert ignore': 'INSERT IGNORE',
            'INSERT IGNORE': 'INSERT IGNORE',
            3: 'REPLACE',
            "replace": 'REPLACE',
            "REPLACE": 'REPLACE',
        }.get(mode, mode)

    def get_count(self, index, *args, **kwargs):
        count = get_line_num_fast(self.gen_path_by_index(index))
        if count:
            count -= 1
        return count

    def get_data(self, index):
        with open(self.gen_path_by_index(index), 'r', encoding=self.encoding) as f:
            table_info_str = ''
            flag_create_table = False
            table_keys = []

            def is_insert(li):
                return li.lower()[:7] in ('insert ', 'replace')

            def get_table_keys(table_info_str_raw):
                if table_info_str_raw.endswith(';'):
                    table_info_str_raw = table_info_str_raw[-1]
                info_str = f'{table_info_str_raw});'
                return list(SqlExtractor(info_str).get_v_main()[0].keys())

            for line in f:
                line = line.strip()
                if line.lower().startswith('create table'):
                    flag_create_table = True
                elif line.lower() == '-- ----------------------------':
                    flag_create_table = False
                    if table_info_str and not table_keys:
                        table_keys = get_table_keys(table_info_str)
                elif is_insert(line):
                    flag_create_table = False
                    if table_info_str and not table_keys:
                        table_keys = get_table_keys(table_info_str)
                if flag_create_table:
                    if line.strip().startswith('`') or line.strip().lower().startswith('create table'):
                        table_info_str += line
                if is_insert(line):
                    # for data in SqlExtractor(line.strip()).get_v_main(table_keys):
                    #     yield data
                    tbname, ks, vss = match_insert(line)
                    for vs in vss:
                        yield {k: v for k, v in zip(ks or table_keys, vs)}

        # con = sqlite3.connect(f'{self.gen_path_by_index(index)}')
        # query = f'select * from {index}'
        # data = pd.read_sql(query, con)
        # for data in data.to_dict():
        #     print(data)
        #     yield data
        # con.close()

    def save_data(self, index, data, update=None, *args, **kwargs):
        if self.compress:
            d_keys = ', '.join('`{}`'.format(data[0].keys()))
            self._file_w[index].write(
                '{} INTO `{}`({}) VALUES ({}){};\n'.format(
                    self.mode,
                    index,
                    d_keys,
                    '), ('.join(
                        ', '.join(
                            v.__str__() if isinstance(v, int) else (
                                'NULL' if isinstance(v, type(None)) else (
                                    "'{}'".format(
                                        v.__str__().replace(
                                            '\\"', RANDOM_STR).replace(
                                            "\\", '\\\\').replace(
                                            "'", '\\\'').replace(
                                            RANDOM_STR, '\\"')) if isinstance(v, str) else (
                                        f"{v.__str__().__repr__()}"
                                    )
                                )
                            )
                            for v in d.values()
                        )
                        for d in data
                    ),
                    f' {update}' if update else ''
                )
            )
        else:
            self._file_w[index].writelines(
                '{} INTO `{}`({}) VALUES ({}){};\n'.format(
                    self.mode,
                    index,
                    ', '.join('`{}`'.format(dk) for dk in d.keys()),
                    ', '.join(
                        v.__str__() if isinstance(v, int) else (
                            'NULL' if isinstance(v, type(None)) else (
                                "'{}'".format(v.__str__().replace(
                                    '\\"', RANDOM_STR).replace(
                                    "\\", '\\\\').replace(
                                    "'", '\\\'').replace(
                                    RANDOM_STR, '\\"')
                                ) if isinstance(v, str) else (
                                    f"{v.__str__().__repr__()}"
                                )
                            ))
                        for v in d.values()
                    ),
                    f' {update}' if update else ''
                )
                for d in data)

        self._file_w[index].flush()

    def create_index(self, index, data, pks='id'):
        super(self.__class__, self).create_index(index, data)
        # self._file_w[index].writerow((self.split.join(f'"{v.__repr__()[1:-1]}"' for v in data.keys()) + '\n'))


class JsonListD(BaseFileD):
    def __init__(self, path, extension='json', encoding='utf8'):
        super().__init__(path, extension, encoding)

    def get_data(self, index):
        with open(self.gen_path_by_index(index), 'r', encoding=self.encoding) as f:
            for line in f:
                yield json.loads(line.strip())

    def save_data(self, index, data, *args, **kwargs):
        self._file_w[index].writelines((json.dumps(d) + '\n') for d in data)
        self._file_w[index].flush()


class XlsIbyFileD(BaseFileD):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    def __init__(self, path, extension='xls'):
        self.path = path
        self.extension = extension
        self._file_w = {}
        self._file_line_num = {}
        self._file_w_now_sheet = {}
        self._file_w_now_sheet_num = {}
        self._file_w_now_sheet_line_num = {}
        self._file_w_now_keys = {}
        self._indexes_path = {}
        super().__init__(path, extension)

    def get_count(self, index, *args, **kwargs):
        all_line = 0
        workbook = xlrd.open_workbook(self.gen_path_by_index(index))
        for idx, name in enumerate(workbook.sheet_names()):
            logging.info(f'sheet:{idx}:{name}')
            worksheet = workbook.sheet_by_index(idx)
            all_line += worksheet.nrows
        return all_line

    def get_data(self, index):
        workbook = xlrd.open_workbook(self.gen_path_by_index(index))  # 文件路径
        # 获取所有sheet的名字
        for idx, name in enumerate(workbook.sheet_names()):
            logging.info(f'sheet:{idx}:{name}')
            worksheet = workbook.sheet_by_index(idx)
            nrows = worksheet.nrows
            if not nrows:
                continue
            keys = {i: key for i, key in enumerate(worksheet.row_values(0))}
            for line_num in range(1, nrows):
                line = worksheet.row_values(line_num)
                # shengri = worksheet.row(line_num)[3].ctype
                # print(shengri)
                data = {keys[i]: key for i, key in enumerate(line)}
                yield data

    def save_data(self, index, data, *args, **kwargs):
        for d in data:
            if not d:
                continue
            row = [str(d.get(key, '')) for key in self._file_w_now_keys[index]]
            self.write_row(self._file_w_now_sheet[index], row)
            self._file_w_now_sheet_line_num[index] += 1
            if self._file_w_now_sheet_line_num[index] > 500000:
                self._file_w_now_sheet_num[index] += 1
                self._file_w_now_sheet[index] = self._file_w[index].create_sheet(
                    index=self._file_w_now_sheet_num[index])
                keys = [key for key in d.keys()]
                self.write_row(self._file_w_now_sheet[index], keys)
                self._file_w_now_sheet_line_num[index] = 1
                self._file_w_now_keys[index] = keys

    @classmethod
    def write_row(cls, ws, row, indices=None):
        try:
            ws.append(row)
        except openpyxl.utils.exceptions.IllegalCharacterError as ex:
            if not indices:
                logging.warning('Failed to write excel: {}'.format(ex))
                return
            for index in indices:
                row[index] = re.sub(cls.ILLEGAL_CHARACTERS_RE, '', row[index])
            try:
                ws.append(row)
            except Exception as ex:
                logging.warning('Failed to write excel: {}\n{}'.format(ex, traceback.format_exc()))

    @classmethod
    def w_open_func(cls, *args, **kwargs):
        wb = openpyxl.Workbook()
        return wb

    def create_index(self, index, data, pks='id'):
        super().create_index(index, data)
        self._file_w_now_sheet_num[index] = 0
        self._file_w_now_sheet[index] = self._file_w[index].create_sheet(index=self._file_w_now_sheet_num[index])
        keys = [key for key in data.keys()]
        self.write_row(self._file_w_now_sheet[index], keys)
        self._file_w_now_sheet_line_num[index] = 1
        self._file_w_now_keys[index] = keys

    def __del__(self):
        for idx, f in self._file_w.items():
            while True:
                try:
                    start_time = time.time()
                    f.save(self._indexes_path[idx])
                    logging.info(
                        '[use {:.2f}s] write excel fin: {}'.format(time.time() - start_time, self._indexes_path[idx]))
                    break
                except ImportError as e:
                    break
                except Exception as e:
                    logging.error(e)
                    traceback.print_exc()
                    time.sleep(5)


class XlsxIbyFileD(XlsIbyFileD):
    def __init__(self, path, extension='xlsx'):
        super().__init__(path=path, extension=extension)


class MongoDBD(object):
    def __init__(self, hosts="mongodb://localhost:27017/", database='test', batch_size=1000):
        self.hosts = hosts
        self.batch_size = batch_size
        self.database = database
        self.client = pymongo.MongoClient(self.hosts)
        self.db_list = self.client.list_database_names()
        self.db = self.client[self.database]
        # self.db.authenticate(user, password)
        self.collection_list = self.db.list_collection_names()
        self.gridfs = gridfs.GridFS(self.db)

    def __repr__(self):
        return f'MongoDB:{self.hosts}/{self.database}'

    def get_data(self, index, *args, **kwargs):
        for d in self.db[index].find().batch_size(self.batch_size):
            d.pop('_id')
            yield d
            # print(list(type(v) for k, v in d.items()))
        #     ObjectId
        # return

    def save_data(self, index, data, *args, **kwargs):
        res = self.db[index].insert_many(data)
        return res.inserted_ids

    def get_indexes(self):
        return list(self.db.list_collection_names())

    def get_count(self, index, *args, **kwargs):
        return self.db[index].find().count()

    def create_index(self, index, data, pks='id'):
        # sql = """SET NAMES utf8mb4;"""
        try:
            self.db[index].rename(f'{index}_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}_bak')
        except pymongo.errors.OperationFailure as e:
            if e.__str__() != 'source namespace does not exist':
                raise

    def gridfs_save_file(self, content):
        with open(content, 'rb') as f:
            data = f.read()
            return self.gridfs.put(data)

    def gridfs_get_file(self, id):
        gf = self.gridfs.get(ObjectId(id))
        return gf.read()

    def gridfs_get_filename(self):
        return self.gridfs.list()


class BaseClient(ABC):
    def __init__(self):
        self.host = None
        self.port = None
        self.database = None
        self.username = None
        self._conn = None
        self.cur_lock = Lock()

    @abc.abstractmethod
    def connect(self):
        raise NotImplementedError('TODO connect function')

    def reconnect(self):
        self._conn = self.connect()

    @property
    def conn(self):
        if not self._conn:
            self.reconnect()
        return self._conn

    @property
    def cur(self):
        return self.conn.cursor()

    @classmethod
    def gen_insert_sql(cls, table_name, data, duplicate_update=False, partition='',):
        keys = []
        values = []
        for k, v in data.items():
            keys.append(k)
            if isinstance(v, (tuple, set)):
                v = list(v)
            if isinstance(v, (dict, list)):
                v = json.dumps(v)
            values.append(v)
        sql = f'insert into {table_name} {partition} ({",".join("`{}`".format(k) for k in keys)}) ' \
              f'values({",".join(["%s"] * len(values))})' + (
                  f'ON DUPLICATE KEY UPDATE {",".join("%s=(%s)" % (k, k) for k in keys)}'
                  if duplicate_update else '')
        return sql, values

    def insert_data(self, table_name, data, duplicate_update=False, partition=''):
        sql, values = self.gen_insert_sql(table_name, data, duplicate_update, partition)
        return self.execute(sql, parms=values)

    def insert_many_by_values_list(self, sql, insert_rows: [], batch_size=4000):
        insert_num = int((len(insert_rows) + batch_size - 1) / batch_size)
        for i in range(insert_num):
            temp = insert_rows[i * batch_size:(i + 1) * batch_size:]
            logging.info('准备插入{}条数据'.format(len(temp)))
            try:
                self.executemany(sql, temp)
            except Exception as ex:
                logging.error(format_error(ex))

    def insert_many_by_data(self, table_name, insert_rows: [], batch_size=4000, duplicate_update=False, partition=''):
        if not insert_rows:
            return
        sql, values = self.gen_insert_sql(table_name, insert_rows[0], duplicate_update, partition)
        insert_num = int((len(insert_rows) + batch_size - 1) / batch_size)
        for i in range(insert_num):
            temp = insert_rows[i * batch_size:(i + 1) * batch_size:]
            logging.info('准备插入{}条数据'.format(len(temp)))
            try:
                self.executemany(sql, [tuple(v.values()) for v in temp])
            except Exception as ex:
                logging.error(format_error(ex))
                raise ex

    @classmethod
    def gen_partition_sql(cls, **kwargs):
        return "partition({})".format(', '.join(f'{k}={v.__repr__()}' for k, v in kwargs.items()))

    def insert_random_data(self, table_name):
        keys, types = self.get_tables_struct(table_name)
        values = [gen_pass(v) for v in types]
        data = {k: v for k, v in zip(keys, values)}
        self.insert_data(table_name, data)
        # self.insert_data_from_table_orc(table_name, data, tables_struct={k: t for k, t in zip(keys, types)})

    def insert_data_from_table_orc(self, table_name, data, tables_struct=None):
        tmp_table_name = self.create_tmp_table(table_name, tables_struct)
        self.insert_data(tmp_table_name, data)
        self.execute(f'INSERT INTO TABLE {table_name} SELECT * FROM {tmp_table_name}')
        self.drop_tmp_table(tmp_table_name)

    def create_table_group_users(self):
        data = {
            "id": 'bigint',
            "chat_id": 'bigint',
            # "group_id": 'bigint',
            "user_id": 'bigint',
            "timestamp": 'bigint',
            "identity": 'INT',
            "last_active_time": 'bigint',
            "in_group": 'INT'
        }
        self.create_table('group_users', data, partitioned={
            "group_id": "string",
            # "year": "int",
            # "month": "int",
        })
        # self.create_table('group_users_no_orc', data, stored='')

    def create_table(
            self, table_name, data,
            partitioned=None,
            # stored='''STORED AS ORC'''
            stored='''STORED AS parquet'''
            # COLLECTION ITEMS TERMINATED BY '\t'
            # MAP KEYS TERMINATED BY '='
    ):
        create_sql = f"""
        CREATE TABLE IF NOT EXISTS `{table_name}`  (
            {', '.join('`%s` %s' % (k, v) for k, v in data.items())}
        )
        {"partitioned by ({})".format(','.join(
            '%s %s' % (k, v) for k, v in partitioned.items())) if partitioned else ''}
        row format delimited fields terminated by '\001'  --csv分隔
        collection items terminated by '\002'  --数组分隔
        map keys terminated by '\003'  --map分隔
        lines terminated by '\n'  --换行分隔
        {'tblproperties("skip.header.line.count"="0")    --跳过文件行首0行' if not stored else ''}
        {stored}
        """
        print(create_sql)
        self.execute(create_sql)

    def drop_tmp_table(self, table_name):
        if table_name.find('no_orc') > 0:
            print(f'DROP TABLE {table_name}')
            self.execute(f'DROP TABLE {table_name}')
            print(f'ok  DROP TABLE {table_name}')

    def create_tmp_table(self, table_name, tables_struct=None):
        if not tables_struct:
            tables_struct = {k: v for k, v in zip(*self.get_tables_struct(table_name))}
        tmp_table_name = f'{table_name}_no_orc_tmp_{time.time()}'.replace(".", "_")
        self.create_table(tmp_table_name, tables_struct, stored='')
        return tmp_table_name

    def load_data(self, path, table_name, tmp_table=False, local=False, tables_struct=None, partition=''):
        # 将临时表中的数据导入到ORC表中, 建立临时表1s load临时表1s 删除临时表1s
        if tmp_table:
            table_name_tmp_table = self.create_tmp_table(table_name, tables_struct)
            try:
                # "load data inpath '/user/admin/data/titanic/train.csv' OVERWRITE into table tmp"
                self.load_data(
                    path, table_name_tmp_table,
                    tmp_table=False, local=local, tables_struct=tables_struct, partition=partition
                )
                self.execute(f"insert into table {table_name} {partition} select * from {table_name_tmp_table}")
            except Exception as e:
                logging.error(format_error(e))
            finally:
                self.drop_tmp_table(table_name_tmp_table)
        else:
            self.execute(f"load data{' local' if local else ''} inpath '{path}' into table {table_name} {partition}")

    @abc.abstractmethod
    def show_table_struct(self, table_name):
        """
        查看表结构
        """
        raise NotImplementedError('TODO show_table_struct function')

    # @abc.abstractmethod
    # def show_table_info(self, table_name):
    #     """
    #     查看表详细属性
    #     """
    #     raise NotImplementedError('TODO show_table_info function')

    @abc.abstractmethod
    def get_tables_struct(self, table_name):
        """
        获取表详细属性list
        """
        raise NotImplementedError('TODO get_tables_struct function')

    def show_tables(self):
        return self.execute('SHOW tables')

    @with_cur_lock()
    def executemany(self, cur, sql, parms=None, *args, **kwargs):
        cur.executemany(sql, parms, *args, **kwargs)
        return cur.fetchall()

    @with_cur_lock()
    def execute(self, cur, sql, parms=None, *args, **kwargs):
        print(sql)
        cur.execute(sql, parms, *args, **kwargs)
        return cur.fetchall()

    @with_cur_lock()
    def execute_iter(self, cur, sql, parms=None, *args, **kwargs):
        print(sql)
        cur.execute(sql, parms, *args, **kwargs)
        result = cur.fetchmany(1000)
        while result:
            for r in result:
                yield r
            result = cur.fetchmany(1000)


class ClickHouseD(BaseClient):
    min_datetime = datetime.datetime(1971, 1, 1, 0, 0, 0)
    max_datetime = datetime.datetime(2106, 1, 1, 0, 0, 0)
    default_types = {'String': '', 'Int': 0, 'Float': 0.0, 'DateTime': min_datetime}

    def __init__(self, host, port, user, password, database):
        super().__init__()
        self.host = host
        self.port = port
        self.user = user
        self.password = password
        self.database = database

    def __repr__(self):
        return f'ClickHouse:{self.host}'

    def connect(self):
        return clickhouse_connect(
            user=self.user,
            password=self.password,
            host=self.host,
            port=self.port,
            database=self.database,
        )

    def get_tables_struct(self, table_name):
        vs_info = self.show_table_struct(table_name)
        keys = [v[0] for v in vs_info]
        types = [v[1] for v in vs_info]
        return keys, types

    def show_table_struct(self, table_name):
        """
        查看表结构
        """
        return self.execute(f'desc {table_name}')

    def cols_type(self, table_name):
        return {kv[0]: kv[1] for kv in self.show_table_struct(table_name)}

    def get_count(self, table_name):
        return self.execute(f'select count(*) from {table_name}')[0][0]

    @classmethod
    def gen_insert_sql_csv(cls, table_name, data):
        keys = []
        values = []
        for k, v in data.items():
            keys.append(k)
            if isinstance(v, (tuple, set)):
                v = list(v)
            if isinstance(v, (dict, list)):
                v = json.dumps(v)
            values.append(v)
        sql = f'insert into {table_name} select {",".join(v.__repr__() for v in values)}'
        return sql, values

    @classmethod
    def gen_insert_sql_no_v(cls, table_name, data):
        keys = []
        values = []
        for k, v in data.items():
            keys.append(k)
            if isinstance(v, (tuple, set)):
                v = list(v)
            if isinstance(v, (dict, list)):
                v = json.dumps(v)
            values.append(v)
        sql = f'INSERT INTO {table_name} ({",".join("`{}`".format(k) for k in keys)}) VALUES'
        return sql, values

    @classmethod
    def format_data_range(cls, data):
        for k, v in data.items():
            if isinstance(data[k], datetime.date):
                if data[k] < ClickHouseD.min_datetime or data[k] > ClickHouseD.max_datetime:
                    data[k] = ClickHouseD.min_datetime
                if getattr(data[k], 'read', None):
                    data[k] = data[k].read()
        return data

    def data_types_check(self, data, cols_types):
        """table_from和table_to数据类型不一致时使用，会降低性能, 默认False"""
        type_mapping = {'String': str, 'Float': float, 'Int': int}
        for k, v in data.items():
            if type(v).__name__ not in cols_types[k].lower():
                for type_k in type_mapping:
                    if cols_types[k].find(type_k) != -1:
                        try:
                            data[k] = type_mapping[type_k](v)
                        except Exception as e:
                            print(type(e), e)
                            data[k] = self.default_types[type_k]
        return data

    def save_data(self, index, data, windows=10000, *args, **kwargs):
        if not data:
            return
        insert_rows = [_ for _ in data if _ and any(_.values())]
        if not insert_rows:
            return
        sql, values = self.gen_insert_sql_no_v(index, insert_rows[0])
        insert_num = int((len(insert_rows) + windows - 1) / windows)
        for i in range(insert_num):
            temp = insert_rows[i * windows:(i + 1) * windows:]
            logging.info('准备插入{}条数据'.format(len(temp)))
            try:
                self.executemany(
                    sql,
                    [self.format_data_range(v) for v in temp])
            except Exception as ex:
                logging.error(format_error(ex))
                raise ex

    def insert_from_mysql(self, table_to, host, port, database, table, user, password):
        # 直接插入到clickhouse现有表
        sql = f"""insert into {table_to} SELECT * FROM mysql('{host}:{port}', '{database}', '{table}', '{user}', '{password}')"""
        self.execute(sql, types_check=True)

    def is_table_exist(self, index):
        try:
            if self.execute(f'select count(*) from {index}'):
                return True
        except Exception:
            return False

    def create_index(self, index, data, pks='id'):
        if self.is_table_exist(index):
            return
        sql = """create table if not exists {}(""".format(index)
        for name, value in data.items():
            if isinstance(value, int):
                _type = 'Int64'
            elif isinstance(value, str) or value is None:
                _type = 'String'
            elif isinstance(value, datetime.date):
                _type = 'DateTime'
            elif isinstance(value, float):
                _type = 'Float'
            else:
                print(value, type(value))
                raise

            sql += ('`' + name + '` ')
            sql += (_type + ',')

        sql = sql.strip(',') + ')'

        sql += """ENGINE = Memory();"""
        self.execute(sql)

    def set_default_data(self, index):
        ch_struct = self.get_tables_struct(index)
        default_value = []
        for col_type in ch_struct[1]:
            d_len = len(default_value)
            for t in self.default_types:
                if t in col_type:
                    default_value.append(self.default_types[t])
                    break
            if len(default_value) != d_len + 1:
                print(col_type)
                raise
        default_data = dict(zip(ch_struct[0], default_value))
        return default_data

    def show_create_table(self, index, distribution=False):
        if distribution:
            pass
        else:
            return self.execute(f'show create table {index}')[0][0].replace('\n', ' ')

    def get_data(self, index, *args, **kwargs):
        sub_sql = f"{index} {kwargs['condition']}" if 'condition' in kwargs else index
        if index.lower().strip().startswith('select '):
            return self.execute_iter(sql=f'{sub_sql}', *args, **kwargs)
        else:
            return self.execute_iter(sql=f'select * from {sub_sql}', *args, **kwargs)


class ListD:
    def __init__(self, index, data=None):
        self.data = {index: data or []}

    def __repr__(self):
        return f'ListD:{list(self.data.keys())}'

    def get_data(self, index, *args, **kwargs):
        return self.data[index]

    def save_data(self, index, data, *args, **kwargs):
        self.data[index].extend(data)

    def get_indexes(self):
        return list(self.data.keys())

    def get_count(self, index, *args, **kwargs):
        return len(self.data[index])


class OracleD(BaseClient):

    def __init__(self, user, password, dsn, database):
        super().__init__()
        self.user = user
        self.password = password
        self.dsn = dsn
        self.database = database

    def __repr__(self):
        return f'Oracle:{self.dsn}'

    def connect(self):
        from cx_Oracle import connect as oracle_connect
        return oracle_connect(
            user=self.user,
            password=self.password,
            dsn=self.dsn,
            encoding='UTF-8'
        )

    @with_cur_lock()
    def execute(self, cur, sql, batch=100000, *args, **kwargs):
        print(sql)
        cur.execute(sql, *args, **kwargs)
        return cur.fetchall()

    def get_column_name(self, index):
        sql = f"select column_name from user_tab_columns  where Table_Name='{index.upper()}'"
        result = ()
        for c in self.execute(sql):
            result = result + c
        return result

    def get_column_name1(self, index):
        sql = f"select column_name from all_tab_cols where table_name='{index.upper()}'"
        return [c[0] for c in self.execute(sql)]

    def get_count(self, index, *args, **kwargs):
        sql = f'select count(*) from {self.database}.{index.upper()}'
        if 'condition' in kwargs:
            sql += kwargs['condition']
        return self.execute(sql)[0][0]

    @with_cur_lock()
    def _execute(self, cur, sql, values=None, batch=10000, **kwargs):
        from cx_Oracle import LOB as oracle_LOB, DatabaseError as oracle_DatabaseError
        print(sql)
        try:
            if values:
                cur.execute(sql, values or [])
            else:
                cur.execute(sql)

            def _fetch(_dbcur):
                try:
                    result = _dbcur.fetchmany(batch)
                except oracle_DatabaseError:
                    time.sleep(600)
                    result = run_task_auto_retry(_dbcur.fetchmany, kwargs={'batch': batch})
                while result:
                    for r in result:
                        yield r
                    try:
                        result = _dbcur.fetchmany(batch)
                    except oracle_DatabaseError:
                        time.sleep(600)
                        result = run_task_auto_retry(_dbcur.fetchmany, kwargs={'batch': batch})
            return _fetch(cur)
        except Exception as e:
            raise e

    def get_data(self, index, *args, **kwargs):
        cols = self.get_column_name1(index)
        sql = f"select {','.join(cols)} from {self.database}.{index}"
        if 'condition' in kwargs:
            sql += kwargs['condition']
        result = self._execute(sql)
        for r in result:
            yield dict(zip(tuple(cols), r))

    def get_tables_struct(self, table_name):
        vs_info = self.show_table_struct(table_name)
        keys = [v[0] for v in vs_info]
        types = [v[1] for v in vs_info]
        return keys, types

    def show_table_struct(self, table_name):
        """
        查看表结构
        """
        return self.execute(f'desc {table_name}')
