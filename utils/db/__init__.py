#!/usr/bin/env python3
# coding: utf-8


import os
import random
import re
import time
import traceback
import openpyxl.utils.exceptions
import openpyxl
import xlrd
import json
import logging
import datetime
from abc import ABC
from pathlib import Path
from .myutils import EsModel
from .myutils import BaseDB
from .myutils import ClientPyMySQL
from .myutils import ExcelWriter


def get_realpath():
    return os.path.split(os.path.realpath(__file__))[0]


class ElasticSearchD(EsModel):
    def __init__(self, hosts, ik=False):
        super().__init__(hosts)
        self.hosts = hosts
        self.ik = ik

    def __repr__(self):
        return f'ElasticSearch:{self.hosts}'

    def get_data(self, index, *args, **kwargs):
        for i in self.scan(query={}, index=index, *args, **kwargs):
            r = {}
            if r.get('id'):
                r['_id'] = i['_id']
            else:
                r['id'] = i['_id']
            r.update(i['_source'])
            yield r

    def save_data(self, index, data, batch_size=1000, uuid_key=('id',), *args, **kwargs):
        actions = []
        for d in data:
            _id = '-'.join(d[k] for k in uuid_key)
            actions.append({
                '_op_type': 'index',  # 操作 index update create delete
                '_index': index,  # index
                '_type': '_doc',  # type
                "_id": f"{_id}",
                '_source': d})
            if len(actions) > batch_size:
                self.bulk_write(actions, *args, **kwargs)
                actions = []

        if len(actions):
            self.bulk_write(actions, *args, **kwargs)

    def get_indexes(self):
        return list(self.es.indices.get_alias().keys())

    def get_count(self, index):
        return int(self.es.count(index=index, body={})['count'])

    @staticmethod
    def get_int_type_from_len(length):
        if 0 < length <= 8:
            return "byte"
        elif 8 < length <= 16:
            return "short"
        elif 16 < length <= 32:
            return "integer"
        elif 32 < length <= 64:
            return "long"
        return None

    @staticmethod
    def get_str_type_from_len(length):
        if 0 < length <= 256:
            return "keyword"
        elif 256 < length <= 65536:
            return "text"
        elif 65536 < length <= 16777216:
            return "text"
        elif 16777216 < length <= 4294967295:
            return "text"
        return None

    def create_index(self, index, data, pks='id'):
        # sql = """SET NAMES utf8mb4;"""
        # sql += """SET FOREIGN_KEY_CHECKS = 0;"""
        # if drop == 1:
        #     sql += """DROP TABLE IF EXISTS `{}`;""".format(tbname)
        pairs = {
            "settings": {
                "index": {
                    "number_of_shards": 12,
                    "number_of_replicas": 1
                }
            },
            'mappings': {
                'properties': {

                }
            }
        }
        properties = pairs['mappings']['properties']

        for name, value in data.items():
            properties[name] = {}
            if isinstance(value, int):
                properties[name]['type'] = 'long'
            elif isinstance(value, str):
                properties[name] = {
                    "type": "text",
                    "fields": {
                        "keyword": {
                            "type": "keyword",
                            "ignore_above": 1024
                        }
                    },
                }
                if self.ik:
                    properties[name]["fields"]["analyzer"] = "ik_max_word"
                    properties[name]["fields"]["search_analyzer"] = "ik_smart"
        if self.es.indices.exists(index=index) is not True:
            res = self.es.indices.create(index=index, body=pairs)
            if not res:
                logging.error("错误，创建索引{}失败".format(index))
                return False

            logging.info("正常，创建索引{}成功".format(index))
            return True
        else:
            logging.info("正常，索引{}已存在".format(index))
            return True


# class MySqlD(BaseDB, ABC):
#     def __init__(self, host, port, user, passwd, database=None):
#         super().__init__(host, port, database, user, passwd)
#
#     def get_data(self, index, *args, **kwargs):
#         return self.__execute__(sql=f'select * from {index}', *args, **kwargs).fetchall()
#
#     def save_data(self, index, *args, **kwargs):
#         self.insert_many_with_dict_list(tablename=index, *args, **kwargs)
#
#     def get_indexes(self):
#         tbs = []
#         for r in list(self.__execute__(sql='show tables;')):
#             tbs.append(r[0])
#         return tbs
#
#     def get_count(self, index):
#         return int(self.__execute__(sql=f'select count(0) as c from {index}').fetchone()[0])
#
#     @staticmethod
#     def get_int_type_from_len(length):
#         if 0 < length <= 8:
#             return "tinyint"
#         elif 8 < length <= 16:
#             return "smallint"
#         elif 16 < length <= 32:
#             return "int"
#         elif 32 < length <= 64:
#             return "bigint"
#         return None
#
#     @staticmethod
#     def get_str_type_from_len(length):
#         if 0 < length <= 256:
#             return "varchar({})".format(length)
#         elif 256 < length <= 65536:
#             return "text"
#         elif 65536 < length <= 16777216:
#             return "mediumtext"
#         elif 16777216 < length <= 4294967295:
#             return "longtext"
#         return None
#
#     def create_index(self, index, data, pks='id'):
#         # sql = """SET NAMES utf8mb4;"""
#         # sql += """SET FOREIGN_KEY_CHECKS = 0;"""
#         # if drop == 1:
#         #     sql += """DROP TABLE IF EXISTS `{}`;""".format(tbname)
#         sql = """create table if not exists {}(""".format(index)
#         for name, value in data.items():
#             if isinstance(value, int):
#                 _type = 'bigint'
#             elif isinstance(value, str):
#                 if pks.find(name) > -1:
#                     _type = 'varchar(256)'
#                 else:
#                     _type = 'text'
#             else:
#                 _type = 'blob'
#
#             sql += ('`' + name + '` ')
#             sql += (_type + ',')
#
#         slen = len(pks)
#         if slen > 0:
#             sql += """PRIMARY KEY ("""
#             v_keys = pks.split(',')
#             first = 1
#             for key in v_keys:
#                 if first == 1:
#                     first = 0
#                     sql += """`{}`""".format(key)
#                 else:
#                     sql += """,`{}`""".format(key)
#             sql += """)) """
#         else:
#             sql = sql[0: -1]
#             sql += ')'
#
#         sql += """ENGINE = InnoDB CHARACTER SET = utf8mb4 COLLATE = utf8mb4_general_ci ROW_FORMAT = Dynamic;"""
#         print(sql)
#         self.__execute__(sql)
#         self.__commit__()


class MySqlD(ClientPyMySQL, ABC):
    def __init__(self, host, port, user, passwd, database=None):
        super().__init__(host, port, database, user, passwd)
        self.host = host
        self.port = port
        self.database = database

    def __repr__(self):
        return f'MySQL:{self.host}:{self.port}/{self.database}'

    def get_data(self, index, *args, **kwargs):
        return self._execute(sql=f'select * from {index}', *args, **kwargs)[1]

    def save_data(self, index, *args, **kwargs):
        self.insert_many_with_dict_list(tablename=index, *args, **kwargs)

    def get_indexes(self):
        return list(t['Tables_in_test'] for t in self._execute(sql='show tables;')[1])

    def get_count(self, index):
        return self._execute(f'select count(0) as c from  {index}', )[1][0]['c']

    @staticmethod
    def get_int_type_from_len(length):
        if 0 < length <= 8:
            return "tinyint"
        elif 8 < length <= 16:
            return "smallint"
        elif 16 < length <= 32:
            return "int"
        elif 32 < length <= 64:
            return "bigint"
        return None

    @staticmethod
    def get_str_type_from_len(length):
        if 0 < length <= 256:
            return "varchar({})".format(length)
        elif 256 < length <= 65536:
            return "text"
        elif 65536 < length <= 16777216:
            return "mediumtext"
        elif 16777216 < length <= 4294967295:
            return "longtext"
        return None

    def create_index(self, index, data, pks='id'):
        # sql = """SET NAMES utf8mb4;"""
        # sql += """SET FOREIGN_KEY_CHECKS = 0;"""
        # if drop == 1:
        #     sql += """DROP TABLE IF EXISTS `{}`;""".format(tbname)
        sql = """create table if not exists {}(""".format(index)
        for name, value in data.items():
            if isinstance(value, int):
                _type = 'bigint'
            elif isinstance(value, str):
                if name in pks.split(','):
                    _type = 'varchar(256)'
                else:
                    _type = 'text'
            else:
                _type = 'blob'

            sql += ('`' + name + '` ')
            sql += (_type + ',')

        slen = len(pks)
        if slen > 0:
            sql += """PRIMARY KEY ("""
            v_keys = pks.split(',')
            first = 1
            for key in v_keys:
                if first == 1:
                    first = 0
                    sql += """`{}`""".format(key)
                else:
                    sql += """,`{}`""".format(key)
            sql += """)) """
        else:
            sql = sql[0: -1]
            sql += ')'

        sql += """ENGINE = InnoDB CHARACTER SET = utf8mb4 COLLATE = utf8mb4_general_ci ROW_FORMAT = Dynamic;"""
        print(sql)
        self._execute(sql)
        self.end_transaction()


def get_line_num_fast(filename):
    count_n = 0
    count_r = 0
    fp = open(filename, "rb")
    while 1:
        buffer = fp.read(1 * 1024 * 1024)
        if not buffer:
            break
        count_n += buffer.count(b'\n')
        count_r += buffer.count(b'\r')
    fp.close()
    return count_n or count_r


class BaseFileD(object):
    def __init__(self, path, extension, encoding='utf8'):
        self.path = path
        self.extension = extension
        self.encoding = encoding
        self._file_w = {}
        self._indexes_path = {}

    def __repr__(self):
        return f'{self.extension}:{self.path}'

    def __del__(self):
        for f in self._file_w.values():
            try:
                f.close()
            except Exception as e:
                logging.warning(e)

    def gen_path_by_index(self, index):
        if index not in self._indexes_path:
            path = f'{self.path}{os.sep}{os.sep.join(index.split("-"))}.{self.extension}'
            self._indexes_path[index] = path
        return self._indexes_path[index]

    def get_indexes(self):
        res = []
        real_path = os.path.split(os.path.realpath(self.path))[0]
        for root, fs, fns in os.walk(self.path):
            for fn in fns:
                path = f'{root}{os.pathsep}{fn}'
                (filepath, tempfilename) = os.path.split(path)
                (filename, extension) = os.path.splitext(tempfilename)
                if extension not in [f'.{self.extension}']:
                    continue
                index = '-'.join(list(filepath.replace(real_path, '', 1).split(os.sep)) + [filename.decode()])
                res.append(index)
                self._indexes_path[index] = path
        return res

    def get_count(self, index):
        return get_line_num_fast(self.gen_path_by_index(index))

    @classmethod
    def w_open_func(cls, *args, **kwargs):
        return open(*args, **kwargs)

    def create_index(self, index, data, pks='id'):
        if not os.path.exists(self.path):
            try:
                os.makedirs(self.path)
            except Exception as e:
                logging.warning(e)
                Path(self.path).mkdir(parents=True, exist_ok=True)
        path = self.gen_path_by_index(index)
        if os.path.exists(path):
            os.rename(path, f"{path}.{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.bak")
        self._file_w.setdefault(index, self.w_open_func(path, 'w', encoding=self.encoding))


class CsvD(BaseFileD):
    def __init__(self, path, split=',', extension='csv', encoding='utf8'):
        super().__init__(path, extension, encoding)
        self.path = path
        self.split = split
        self.extension = extension
        self._file_w = {}
        self._indexes_path = {}

    def get_data(self, index):
        with open(self.gen_path_by_index(index), 'r', encoding=self.encoding) as f:
            keys = list(k.replace("'", '').replace('"', '') for k in f.readline().strip().split(self.split))
            for line in f:
                yield {keys[idx]: v for idx, v in enumerate(eval(line))}

    def save_data(self, index, data, *args, **kwargs):
        self._file_w[index].writelines((','.join(v.__repr__() for v in d.values()) + '\n') for d in data)

    def create_index(self, index, data, pks='id'):
        super(self.__class__, self).create_index(index, data)
        self._file_w[index].write((','.join(v.__str__() for v in data.keys()) + '\n'))


class JsonListD(BaseFileD):
    def __init__(self, path, extension='json', encoding='utf8'):
        super().__init__(path, extension, encoding)
        self.path = path
        self.extension = extension
        self._file_w = {}
        self._indexes_path = {}

    def get_data(self, index):
        with open(self.gen_path_by_index(index), 'r', encoding=self.encoding) as f:
            for line in f:
                yield json.loads(line.strip())

    def save_data(self, index, data, *args, **kwargs):
        self._file_w[index].writelines((json.dumps(d) + '\n') for d in data)


class XlsIbyFileD(BaseFileD):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    def __init__(self, path, extension='xls'):
        self.path = path
        self.extension = extension
        self._file_w = {}
        self._file_line_num = {}
        self._file_w_now_sheet = {}
        self._file_w_now_sheet_num = {}
        self._file_w_now_sheet_line_num = {}
        self._file_w_now_keys = {}
        self._indexes_path = {}
        super().__init__(path, extension)

    def get_count(self, index):
        all_line = 0
        workbook = xlrd.open_workbook(self.gen_path_by_index(index))
        for idx, name in enumerate(workbook.sheet_names()):
            logging.info(f'sheet:{idx}:{name}')
            worksheet = workbook.sheet_by_index(idx)
            all_line += worksheet.nrows
        return all_line

    def get_data(self, index):
        workbook = xlrd.open_workbook(self.gen_path_by_index(index))  # 文件路径
        # 获取所有sheet的名字
        for idx, name in enumerate(workbook.sheet_names()):
            logging.info(f'sheet:{idx}:{name}')
            worksheet = workbook.sheet_by_index(idx)
            nrows = worksheet.nrows
            keys = {i: key.lower() for i, key in enumerate(worksheet.row_values(0))}
            for line_num in range(1, nrows):
                line = worksheet.row_values(line_num)
                data = {keys[i]: key for i, key in enumerate(line)}
                yield data

    def save_data(self, index, data, *args, **kwargs):
        for d in data:
            row = [str(d.get(key, '')) for key in self._file_w_now_keys[index]]
            self.write_row(self._file_w_now_sheet[index], row)
            self._file_w_now_sheet_line_num[index] += 1
            if self._file_w_now_sheet_line_num[index] > 500000:
                self._file_w_now_sheet_num[index] += 1
                self._file_w_now_sheet[index] = self._file_w[index].create_sheet(
                    index=self._file_w_now_sheet_num[index])
                keys = [key for key in d.keys()]
                self.write_row(self._file_w_now_sheet[index], keys)
                self._file_w_now_sheet_line_num[index] = 1
                self._file_w_now_keys[index] = keys

    @classmethod
    def write_row(cls, ws, row, indices=None):
        try:
            ws.append(row)
        except openpyxl.utils.exceptions.IllegalCharacterError as ex:
            if not indices:
                logging.warning('Failed to write excel: {}'.format(ex))
                return
            for index in indices:
                row[index] = re.sub(cls.ILLEGAL_CHARACTERS_RE, '', row[index])
            try:
                ws.append(row)
            except Exception as ex:
                logging.warning('Failed to write excel: {}\n{}'.format(ex, traceback.format_exc()))

    @classmethod
    def w_open_func(cls, *args, **kwargs):
        wb = openpyxl.Workbook()
        return wb

    def create_index(self, index, data, pks='id'):
        super().create_index(index, data)
        self._file_w_now_sheet_num[index] = 0
        self._file_w_now_sheet[index] = self._file_w[index].create_sheet(index=self._file_w_now_sheet_num[index])
        keys = [key for key in data.keys()]
        self.write_row(self._file_w_now_sheet[index], keys)
        self._file_w_now_sheet_line_num[index] = 1
        self._file_w_now_keys[index] = keys

    def __del__(self):
        for idx, f in self._file_w.items():
            while True:
                try:
                    start_time = time.time()
                    f.save(self._indexes_path[idx])
                    logging.info('[use {:.2f}s] write excel fin: {}'.format(time.time()-start_time, self._indexes_path[idx]))
                    break
                except Exception as e:
                    logging.error(e)
                    traceback.print_exc()
                    time.sleep(5)


class XlsxIbyFileD(XlsIbyFileD):
    def __init__(self, path, extension='xlsx'):
        super().__init__(path=path, extension=extension)
